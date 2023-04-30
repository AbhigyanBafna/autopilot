import requests
import openpyxl
from openpyxl import Workbook
import csv
import os
from dotenv import load_dotenv
from config.definitions import MOVIE_DATA, MOVIE_NAMES

load_dotenv()

api_key = os.getenv('OMDB_APIKEY')
endpoint = "http://www.omdbapi.com/"

movies = []
with open(MOVIE_NAMES, newline='') as csvfile:
    reader = csv.reader(csvfile)
    for row in reader:
        movies.extend(row)

wb = openpyxl.load_workbook(filename=MOVIE_DATA)
ws = wb.active

current_row = 2

for movie in movies:
    params = {"apikey": api_key, "t": movie}
    response = requests.get(endpoint, params=params)
    data = response.json()

    if data["Response"] == "False":
        print(f"Error: Could not find movie '{movie}' on OMDb API.")
        continue

    genres = data["Genre"].split(", ")[:2]
    imdb_score = data["imdbRating"]
    if imdb_score is None:
        imdb_score = 5

    ratings = data["Ratings"]
    rt_rating = next((rating["Value"] for rating in ratings if rating["Source"] == "Rotten Tomatoes"), str(50)).replace("%", "")

    if len(genres) > 1:
        ws.cell(row=current_row, column=1, value=movie)
        ws.cell(row=current_row, column=2, value=genres[0])
        ws.cell(row=current_row, column=3, value=genres[1])
        ws.cell(row=current_row, column=4, value=imdb_score)
        ws.cell(row=current_row, column=5, value=rt_rating)
    else:
        ws.cell(row=current_row, column=1, value=movie)
        ws.cell(row=current_row, column=2, value=genres[0])
        ws.cell(row=current_row, column=3, value=genres[0])
        ws.cell(row=current_row, column=4, value=imdb_score)
        ws.cell(row=current_row, column=5, value=rt_rating)

    current_row += 1

wb.save(filename=MOVIE_DATA)